"""Excel export ba openpyxl."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import db
import config


def build_xlsx(path: str = None) -> str:
    path = path or config.XLSX_PATH
    wb = Workbook()
    ws = wb.active
    ws.title = "کازینو"
    ws.sheet_view.rightToLeft = True

    custom = db.list_custom_fields()  # [(cfX, label)]
    headers = (["ردیف"] + [fa for _, fa in db.COLUMNS]
               + [label for _, label in custom] + ["ثبت", "ویرایش"])
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = head_fill

    for i, rec in enumerate(reversed(db.all_records()), start=1):
        extra = db.record_extra(rec)
        row = [i] + [rec.get(k, "") or "" for k, _ in db.COLUMNS]
        row += [extra.get(key, "") or "" for key, _ in custom]
        row += [rec.get("created_at", "") or "", rec.get("updated_at", "") or ""]
        ws.append(row)

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)

    wb.save(path)
    return path
